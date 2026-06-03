import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from src.paths import INTERIM_DIR, ensure_directories
from src.utils.io import read_csv
from src.utils.logging_utils import get_logger


CODING_COLUMNS = [
    "coder_id",
    "human_ai_answerable",
    "basic_howto_debugging",
    "requires_context",
    "sufficient_information",
    "llm_ai_answerable",
    "confidence_1_to_5",
    "notes",
]


VISIBLE_COLUMNS = [
    "question_id",
    "creation_date",
    "tag",
    "all_tags",
    "title",
    "body_excerpt_1500",
    "body_length",
    "has_code",
    "short_code",
    "how_to_error_title",
    "answer_count",
    "has_accepted_answer",
    "is_closed",
    "minutes_to_first_answer",
] + CODING_COLUMNS


INSTRUCTIONS = [
    ["Purpose", "Validate whether the tag-level AI-answerability index reflects actual question-level AI answerability."],
    ["Blind coding", "Do not use tag-level AI scores, strata, or key files while coding."],
    ["Scale", "Use 1 = yes, 0 = no. Leave blank only if impossible to judge."],
    ["human_ai_answerable", "Would a general-purpose generative AI assistant likely produce a useful answer from the question alone?"],
    ["basic_howto_debugging", "Is this a basic how-to, syntax, simple debugging, common error, or short-code task?"],
    ["requires_context", "Does the question require substantial hidden/local/project/version/environment context?"],
    ["sufficient_information", "Does the question include enough information for a knowledgeable assistant to attempt a useful answer?"],
    ["llm_ai_answerable", "Fill only if this file is coded by or with an LLM classification."],
    ["confidence_1_to_5", "1 = very uncertain; 5 = very confident."],
    ["Workflow", "Code all rows in the Coding sheet. Save the completed workbook with coder name/date."],
]


def add_instructions_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.append(["Field", "Instruction"])
    for row in INSTRUCTIONS:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def add_coding_sheet(wb: Workbook, df: pd.DataFrame, coder_id: str) -> None:
    ws = wb.create_sheet("Coding")
    out = df[VISIBLE_COLUMNS].copy()
    out["coder_id"] = coder_id
    ws.append(list(out.columns))
    for row in out.itertuples(index=False, name=None):
        ws.append(list(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 12,
        "B": 20,
        "C": 18,
        "D": 36,
        "E": 55,
        "F": 95,
        "G": 12,
        "H": 10,
        "I": 10,
        "J": 14,
        "K": 12,
        "L": 14,
        "M": 10,
        "N": 16,
        "O": 14,
        "P": 18,
        "Q": 20,
        "R": 16,
        "S": 20,
        "T": 18,
        "U": 18,
        "V": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
        row[5].alignment = Alignment(wrap_text=True, vertical="top")
        row[-1].alignment = Alignment(wrap_text=True, vertical="top")

    binary_validation = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    confidence_validation = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
    ws.add_data_validation(binary_validation)
    ws.add_data_validation(confidence_validation)

    first_row = 2
    last_row = ws.max_row
    for col in ["P", "Q", "R", "S", "T"]:
        binary_validation.add(f"{col}{first_row}:{col}{last_row}")
    confidence_validation.add(f"U{first_row}:U{last_row}")


def add_variable_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Variables")
    rows = [
        ["Variable", "Meaning"],
        ["question_id", "Stack Overflow question identifier."],
        ["title", "Question title."],
        ["body_excerpt_1500", "Plain-text question body excerpt used for coding."],
        ["human_ai_answerable", "1 if answerable by a general-purpose generative AI assistant from the question alone."],
        ["basic_howto_debugging", "1 if basic how-to/simple debugging/common error/short-code task."],
        ["requires_context", "1 if hidden local/project/version/environment context is required."],
        ["sufficient_information", "1 if enough information is present for a useful answer."],
        ["llm_ai_answerable", "Optional LLM-based classification."],
        ["confidence_1_to_5", "Coder confidence, 1 to 5."],
    ]
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="70AD47")
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def build_workbook(input_path: Path, output_path: Path, coder_id: str) -> None:
    df = read_csv(input_path)
    wb = Workbook()
    add_instructions_sheet(wb)
    add_coding_sheet(wb, df, coder_id)
    add_variable_sheet(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Excel workbooks for blind AI-answerability validation coding.")
    parser.add_argument("--input-dir", type=Path, default=INTERIM_DIR / "validation_coding")
    parser.add_argument("--output-dir", type=Path, default=INTERIM_DIR / "validation_coding")
    return parser.parse_args()


def main() -> None:
    ensure_directories()
    args = parse_args()
    logger = get_logger(__name__)
    jobs = [
        ("ai_answerability_validation_full_coder_A.csv", "ai_answerability_validation_full_coder_A.xlsx", "A"),
        ("ai_answerability_validation_full_coder_B.csv", "ai_answerability_validation_full_coder_B.xlsx", "B"),
        ("ai_answerability_validation_pilot_coder_A.csv", "ai_answerability_validation_pilot_coder_A.xlsx", "A"),
        ("ai_answerability_validation_pilot_coder_B.csv", "ai_answerability_validation_pilot_coder_B.xlsx", "B"),
    ]
    for src, dest, coder_id in jobs:
        input_path = args.input_dir / src
        output_path = args.output_dir / dest
        build_workbook(input_path, output_path, coder_id)
        logger.info("Wrote workbook to %s", output_path)


if __name__ == "__main__":
    main()
